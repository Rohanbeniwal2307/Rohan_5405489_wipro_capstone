from pathlib import Path

from openpyxl import load_workbook


class ExcelReader:
    @staticmethod
    def get_rows(file_path, sheet_name=None):
        workbook = load_workbook(Path(file_path))
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return []

        headers = rows[0]
        return [
            dict(zip(headers, row))
            for row in rows[1:]
            if any(value is not None for value in row)
        ]

